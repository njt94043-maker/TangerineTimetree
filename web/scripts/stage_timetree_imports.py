"""
stage_timetree_imports.py — s260 TimeTree migration stager.

Reads the scraped xlsx (both sheets) and UPSERTS into the `import_staging`
table. NOTHING lands on the calendar here — commit is an explicit per-card /
bulk action in the Imports review UI. Re-run safe: committed/skipped decisions
are NEVER modified; only `pending` rows refresh from the sheet.

The scrape is always a FULL calendar export, so this also tracks drift:
  - rows present in the xlsx get last_seen_at stamped (+ missing flags cleared);
  - committed rows whose incoming proposal differs are flagged source_changed;
  - staging rows absent from the xlsx are flagged missing_from_source.

Usage:
    python stage_timetree_imports.py --dry-run     # counts only, no writes
    python stage_timetree_imports.py               # apply
    python stage_timetree_imports.py --xlsx path\to\file.xlsx

Nathan runs this by hand after a scrape — no scheduling, no auto-invocation.
"""
import argparse
import os
import sys
from datetime import datetime, timezone

import openpyxl
import requests

SUPABASE_URL = "https://jlufqgslgjowfaqmqlds.supabase.co"
REST = SUPABASE_URL + "/rest/v1"
DEFAULT_XLSX = r"C:\Apps\timetree-scrape\timetree_gigs_enriched.xlsx"
FALLBACK_XLSX = r"C:\Apps\timetree-scrape\timetree_gigs.xlsx"


def get_service_key() -> str:
    # Dual-path (matches the seed scripts): env var first, DPAPI store fallback.
    v = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if v:
        return v
    sys.path.insert(0, r"C:\apps\Dev Team\scripts")
    from dev_secrets import get_secret  # noqa: E402
    key = get_secret("SUPABASE_SERVICE_ROLE_KEY")
    if not key:
        sys.exit("SUPABASE_SERVICE_ROLE_KEY not found (env or DPAPI store)")
    return key


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def s(v) -> str:
    return "" if v is None else str(v).strip()


def fmt_date(v) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    text = str(v).strip()
    return text[:10] if len(text) >= 10 else text


def to_bool(v) -> bool:
    if isinstance(v, bool):
        return v
    return s(v).lower() in ("true", "1", "yes")


def header_index(ws):
    hdr = [s(c.value).lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    return {name: i for i, name in enumerate(hdr)}


def read_gigs(ws):
    idx = header_index(ws)

    def col(row, name):
        i = idx.get(name)
        return row[i] if i is not None and i < len(row) else None

    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        uid = s(col(row, "timetree uid"))
        if not uid:
            continue
        notes = s(col(row, "notes"))
        out.append({
            "timetree_uid": uid,
            "kind": "gig",
            "action": s(col(row, "action")) or "KEEP",
            "raw_title": s(col(row, "timetree title")),
            "raw_notes": notes,
            "match_source": s(col(row, "match source")),
            "proposed": {
                "date": fmt_date(col(row, "date")),
                "gig_type": (s(col(row, "gig type")).lower() or "gig"),
                "venue": s(col(row, "venue")),
                "client_name": s(col(row, "client name")),
                "fee": s(col(row, "fee")),
                "payment_type": s(col(row, "payment type")),
                "load_time": s(col(row, "load time")),
                "start_time": s(col(row, "start time")),
                "end_time": s(col(row, "end time")),
                "notes": notes,
                "is_public": to_bool(col(row, "public?")),
            },
        })
    return out


def read_aways(ws, profiles):
    idx = header_index(ws)

    def col(row, name):
        i = idx.get(name)
        return row[i] if i is not None and i < len(row) else None

    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        uid = s(col(row, "timetree uid"))
        if not uid:
            continue
        member = s(col(row, "member"))
        out.append({
            "timetree_uid": uid,
            "kind": "away",
            "action": s(col(row, "action")) or "KEEP",
            "raw_title": s(col(row, "timetree title")),
            "raw_notes": "",
            "match_source": "",
            "proposed": {
                "member_name": member,
                "user_id": resolve_member(member, profiles),
                "start_date": fmt_date(col(row, "start date")),
                "end_date": fmt_date(col(row, "end date")),
                "reason": s(col(row, "reason")),
            },
        })
    return out


def resolve_member(first_name, profiles):
    fn = first_name.strip().lower()
    if not fn:
        return None
    for p in profiles:
        if s(p.get("name")).lower().startswith(fn):
            return p["id"]
    return None


def same_proposed(incoming, stored):
    """True if every incoming field matches the stored proposal (over incoming's
    keys — stored-only keys like UI-resolved ids are ignored)."""
    stored = stored or {}
    for k, v in incoming.items():
        if s(v) != s(stored.get(k)):
            return False
    return True


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=None)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    path = args.xlsx
    if not path:
        if os.path.exists(DEFAULT_XLSX):
            path = DEFAULT_XLSX
        elif os.path.exists(FALLBACK_XLSX):
            print(f"WARNING: enriched xlsx not found — falling back to {FALLBACK_XLSX}")
            path = FALLBACK_XLSX
        else:
            sys.exit(f"No xlsx found ({DEFAULT_XLSX} / {FALLBACK_XLSX})")
    if not os.path.exists(path):
        sys.exit(f"xlsx not found: {path}")

    key = get_service_key()
    headers = {"apikey": key, "Authorization": f"Bearer {key}", "Content-Type": "application/json"}

    profiles = requests.get(f"{REST}/profiles?select=id,name", headers=headers, timeout=30).json()
    existing_resp = requests.get(f"{REST}/import_staging?select=*", headers=headers, timeout=30)
    existing_resp.raise_for_status()
    existing = {r["timetree_uid"]: r for r in existing_resp.json()}

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    incoming = read_gigs(wb["TGT Gigs"]) + read_aways(wb["Away Dates"], profiles)
    wb.close()

    incoming_uids = set()
    inserts, patches = [], []  # (row_or_patch) tuples for the real run
    c = dict(new=0, updated=0, protected=0, skipped_junk=0, now_missing=0, restored=0, changed_in_source=0)

    for row in incoming:
        uid = row["timetree_uid"]
        incoming_uids.add(uid)
        ex = existing.get(uid)
        if ex is None:
            # New — junk (action != KEEP) lands as skipped; the rest as pending.
            status = "pending" if row["action"].upper() == "KEEP" else "skipped"
            if status == "skipped":
                c["skipped_junk"] += 1
            else:
                c["new"] += 1
            inserts.append({
                "timetree_uid": uid, "kind": row["kind"], "raw_title": row["raw_title"],
                "raw_notes": row["raw_notes"], "proposed": row["proposed"],
                "match_source": row["match_source"], "status": status, "last_seen_at": now_iso(),
            })
            continue

        patch = {"last_seen_at": now_iso()}
        if ex.get("missing_from_source") or ex.get("missing_acknowledged"):
            patch["missing_from_source"] = False
            patch["missing_acknowledged"] = False
            if ex.get("missing_from_source"):
                c["restored"] += 1

        st = ex.get("status")
        if st == "pending":
            patch.update({
                "proposed": row["proposed"], "match_source": row["match_source"],
                "raw_title": row["raw_title"], "raw_notes": row["raw_notes"],
            })
            c["updated"] += 1
        elif st == "committed":
            if not same_proposed(row["proposed"], ex.get("proposed")):
                patch["source_changed"] = True
                patch["latest_from_source"] = row["proposed"]
                c["changed_in_source"] += 1
            else:
                c["protected"] += 1
        else:  # skipped — decision protected
            c["protected"] += 1
        patches.append((ex["id"], patch))

    # Sweep: staging rows absent from this full export → missing (unless ack'd).
    for uid, ex in existing.items():
        if uid in incoming_uids:
            continue
        if ex.get("missing_acknowledged") or ex.get("missing_from_source"):
            continue
        c["now_missing"] += 1
        patches.append((ex["id"], {"missing_from_source": True}))

    mode = "DRY-RUN (no writes)" if args.dry_run else "APPLIED"
    print(f"--- stage_timetree_imports [{mode}] ---")
    print(f"xlsx: {path}")
    print(f"incoming rows: {len(incoming)}  |  existing staging rows: {len(existing)}")
    print(f"  new (pending):      {c['new']}")
    print(f"  updated (pending):  {c['updated']}")
    print(f"  protected:          {c['protected']}")
    print(f"  skipped-junk:       {c['skipped_junk']}")
    print(f"  changed-in-source:  {c['changed_in_source']}")
    print(f"  now-missing:        {c['now_missing']}")
    print(f"  restored:           {c['restored']}")

    if args.dry_run:
        print("dry-run — nothing written.")
        return

    for rec in inserts:
        r = requests.post(f"{REST}/import_staging", headers={**headers, "Prefer": "return=minimal"},
                          json=rec, timeout=30)
        r.raise_for_status()
    for row_id, patch in patches:
        r = requests.patch(f"{REST}/import_staging?id=eq.{row_id}",
                           headers={**headers, "Prefer": "return=minimal"}, json=patch, timeout=30)
        r.raise_for_status()
    print(f"wrote {len(inserts)} inserts + {len(patches)} patches.")


if __name__ == "__main__":
    main()
