"""
seed_supabase.py — Seed Supabase with gigs + away dates from TimeTree export.

Reads timetree_gigs.xlsx (original) and timetree_gigs_enriched.xlsx (venue/client data).
Only seeds fees that existed in the ORIGINAL timetree calendar.
WhatsApp-enriched fees are stripped — user reviews them separately.

Usage:
    python scripts/seed_supabase.py              # dry run (prints what would be inserted)
    python scripts/seed_supabase.py --execute     # actually insert into Supabase
"""

import os
import sys
import json
import datetime
import argparse

# Add parent for imports
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(SCRIPT_DIR)

# openpyxl + requests
try:
    import openpyxl
    import requests
except ImportError as e:
    print(f"Missing dependency: {e}")
    print("Run: pip install openpyxl requests")
    sys.exit(1)

# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------
SUPABASE_URL = "https://jlufqgslgjowfaqmqlds.supabase.co"
SERVICE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")

# Load from .env if not in environment
ENV_FILE = os.path.join(ROOT_DIR, ".env")
if not SERVICE_KEY and os.path.exists(ENV_FILE):
    with open(ENV_FILE) as f:
        for line in f:
            line = line.strip()
            if line.startswith("SUPABASE_SERVICE_ROLE_KEY="):
                SERVICE_KEY = line.split("=", 1)[1]

if not SERVICE_KEY:
    print("ERROR: SUPABASE_SERVICE_ROLE_KEY not found in .env or environment")
    sys.exit(1)

HEADERS = {
    "apikey": SERVICE_KEY,
    "Authorization": f"Bearer {SERVICE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=representation",
}

# Band member name → Supabase profile UUID
MEMBER_IDS = {
    "Nathan": "f30962b3-2588-4b3d-827a-69b03bdfa6b1",
    "Neil": "8d6993a3-444d-4c8f-8a8a-fac05a3d06be",
    "James": "2c856d6c-1ddb-464f-9db3-15e5ca463595",
    "Adam": "43b1dc8e-9258-4266-be2c-5052d5a0434c",
}

# Default creator for seeded gigs (Nathan = admin)
DEFAULT_CREATOR = MEMBER_IDS["Nathan"]

# Data files
SCRAPE_DIR = os.path.join(os.path.dirname(ROOT_DIR), "timetree-scrape")
ORIGINAL_FILE = os.path.join(SCRAPE_DIR, "timetree_gigs.xlsx")
ENRICHED_FILE = os.path.join(SCRAPE_DIR, "timetree_gigs_enriched.xlsx")


# ---------------------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------------------
def fmt_date(val):
    """Convert datetime to YYYY-MM-DD string."""
    if isinstance(val, datetime.datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, datetime.date):
        return val.strftime("%Y-%m-%d")
    return str(val) if val else None


def fmt_time(val):
    """Convert time value to HH:MM string or None."""
    if not val:
        return None
    s = str(val).strip()
    if len(s) >= 5 and ":" in s:
        return s[:5]
    return None


def parse_fee(val):
    """Parse fee to float or None."""
    if not val:
        return None
    s = str(val).strip().replace("£", "").replace(",", "")
    try:
        f = float(s)
        return f if f > 0 else None
    except ValueError:
        return None


def supabase_insert(table, rows, execute=False):
    """Insert rows into Supabase table. Returns (inserted_count, errors)."""
    if not execute:
        return len(rows), []

    errors = []
    inserted = 0

    # Batch insert (Supabase REST API supports array of objects)
    BATCH_SIZE = 50
    for i in range(0, len(rows), BATCH_SIZE):
        batch = rows[i : i + BATCH_SIZE]
        r = requests.post(
            f"{SUPABASE_URL}/rest/v1/{table}",
            headers=HEADERS,
            json=batch,
        )
        if r.status_code in (200, 201):
            result = r.json()
            inserted += len(result) if isinstance(result, list) else 1
        else:
            errors.append(f"Batch {i//BATCH_SIZE + 1}: {r.status_code} {r.text[:200]}")

    return inserted, errors


# ---------------------------------------------------------------------------
# LOAD DATA
# ---------------------------------------------------------------------------
def load_gigs():
    """Load gigs from enriched xlsx, but fees only from original."""
    wb_orig = openpyxl.load_workbook(ORIGINAL_FILE)
    ws_orig = wb_orig["TGT Gigs"]

    wb_enr = openpyxl.load_workbook(ENRICHED_FILE)
    ws_enr = wb_enr["TGT Gigs"]

    gigs = []
    whatsapp_fees = []  # For user review

    for row in range(2, ws_enr.max_row + 1):
        action = ws_enr.cell(row=row, column=1).value
        if action and str(action).strip().upper() == "DELETE":
            continue

        date_val = ws_enr.cell(row=row, column=3).value
        if not date_val:
            continue

        title = ws_enr.cell(row=row, column=2).value or ""
        gig_type = ws_enr.cell(row=row, column=4).value or "gig"
        venue = ws_enr.cell(row=row, column=5).value or ""
        client = ws_enr.cell(row=row, column=6).value or ""
        enriched_fee = ws_enr.cell(row=row, column=7).value
        payment_type = ws_enr.cell(row=row, column=8).value or ""
        load_time = ws_enr.cell(row=row, column=9).value
        start_time = ws_enr.cell(row=row, column=10).value
        end_time = ws_enr.cell(row=row, column=11).value
        notes = ws_enr.cell(row=row, column=12).value or ""
        is_public = ws_enr.cell(row=row, column=13).value
        match_source = ws_enr.cell(row=row, column=15).value or ""

        # Original fee (from timetree calendar itself)
        orig_fee = ws_orig.cell(row=row, column=7).value

        # Only use fee if it was in the ORIGINAL timetree export
        fee = parse_fee(orig_fee)

        # Track WhatsApp-enriched fees for user review
        if not orig_fee and enriched_fee and parse_fee(enriched_fee):
            whatsapp_fees.append({
                "date": fmt_date(date_val),
                "title": title,
                "venue": venue,
                "fee": parse_fee(enriched_fee),
            })

        # Clean up venue from title if no enriched venue
        if not venue and title:
            import re
            cleaned = re.sub(r"^Gig\s*[-\u2013]?\s*", "", title.strip(), flags=re.IGNORECASE)
            cleaned = cleaned.strip(" -()")
            if cleaned:
                venue = cleaned

        # Normalize gig_type
        gig_type = gig_type.strip().lower() if gig_type else "gig"
        if gig_type not in ("gig", "practice"):
            gig_type = "gig"

        # Normalize payment_type
        pt = str(payment_type).strip().lower() if payment_type else ""
        if pt == "agency":
            pt = "invoice"
        if pt not in ("cash", "invoice"):
            pt = ""

        # Normalize is_public
        pub = False
        if is_public and str(is_public).strip().upper() == "TRUE":
            pub = True

        date_str = fmt_date(date_val)
        if not date_str:
            continue

        gigs.append({
            "date": date_str,
            "gig_type": gig_type,
            "venue": venue,
            "client_name": client,
            "fee": fee,
            "payment_type": pt,
            "load_time": fmt_time(load_time),
            "start_time": fmt_time(start_time),
            "end_time": fmt_time(end_time),
            "notes": notes,
            "is_public": pub,
            "created_by": DEFAULT_CREATOR,
        })

    return gigs, whatsapp_fees


def load_away_dates():
    """Load away dates from enriched xlsx."""
    wb = openpyxl.load_workbook(ENRICHED_FILE)
    ws = wb["Away Dates"]

    away_dates = []

    for row in range(2, ws.max_row + 1):
        action = ws.cell(row=row, column=1).value
        if action and str(action).strip().upper() == "DELETE":
            continue

        member = ws.cell(row=row, column=2).value or ""
        start_date = ws.cell(row=row, column=3).value
        end_date = ws.cell(row=row, column=4).value
        reason = ws.cell(row=row, column=5).value or ""

        if not member or not start_date:
            continue

        user_id = MEMBER_IDS.get(member.strip())
        if not user_id:
            print(f"  WARNING: Unknown member '{member}' at row {row}, skipping")
            continue

        start_str = fmt_date(start_date)
        end_str = fmt_date(end_date) if end_date else start_str

        if not start_str:
            continue

        away_dates.append({
            "user_id": user_id,
            "start_date": start_str,
            "end_date": end_str,
            "reason": reason,
        })

    return away_dates


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="Seed Supabase from TimeTree export")
    parser.add_argument("--execute", action="store_true", help="Actually insert (default: dry run)")
    args = parser.parse_args()

    print("=" * 70)
    print("TGT Supabase Seed Script")
    print(f"Mode: {'EXECUTE' if args.execute else 'DRY RUN'}")
    print("=" * 70)

    # Check files exist
    for f in [ORIGINAL_FILE, ENRICHED_FILE]:
        if not os.path.exists(f):
            print(f"ERROR: File not found: {f}")
            sys.exit(1)

    # Load data
    print("\nLoading gigs...")
    gigs, whatsapp_fees = load_gigs()
    print(f"  {len(gigs)} gigs to insert")
    gigs_with_fee = [g for g in gigs if g["fee"] is not None]
    print(f"  {len(gigs_with_fee)} with original timetree fees")
    print(f"  {len(whatsapp_fees)} WhatsApp-confirmed fees (NOT seeded — for review)")

    print("\nLoading away dates...")
    away_dates = load_away_dates()
    print(f"  {len(away_dates)} away dates to insert")

    # Print WhatsApp fees for review
    if whatsapp_fees:
        print("\n" + "=" * 70)
        print("WHATSAPP-CONFIRMED FEES (not seeded — double-check these)")
        print("=" * 70)
        for wf in whatsapp_fees:
            print(f"  {wf['date']}  {wf['venue'] or wf['title']:<35}  £{wf['fee']:.0f}")

    # Preview some gigs
    print("\n" + "-" * 70)
    print("SAMPLE GIGS (first 10)")
    print("-" * 70)
    for g in gigs[:10]:
        fee_str = f"£{g['fee']:.0f}" if g["fee"] else "  -"
        print(f"  {g['date']}  {g['venue'][:30]:<30}  {fee_str:>5}  {g['gig_type']}")

    # Insert
    print("\n" + "-" * 70)
    if args.execute:
        print("INSERTING GIGS...")
        count, errs = supabase_insert("gigs", gigs, execute=True)
        print(f"  Inserted: {count}/{len(gigs)}")
        for e in errs:
            print(f"  ERROR: {e}")

        print("\nINSERTING AWAY DATES...")
        count, errs = supabase_insert("away_dates", away_dates, execute=True)
        print(f"  Inserted: {count}/{len(away_dates)}")
        for e in errs:
            print(f"  ERROR: {e}")

        print("\nDone!")
    else:
        print("DRY RUN — no data inserted. Use --execute to insert.")

    print("-" * 70)


if __name__ == "__main__":
    main()
